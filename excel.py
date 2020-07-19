import openpyxl,os,re,time
from datetime import datetime
from tqdm import tqdm

class Excel:
    def get_title_list(self,start_row,end_row):
        self.path=os.path.dirname(__file__)
        self.workbook = openpyxl.load_workbook(self.path+'/'+'Data_2020_07_13.xlsx')#读入文件的路径 
        self.booksheet=self.workbook.active
        max_row = self.booksheet.max_row
        self.min_row=start_row#开始
        self.max_row=end_row#结束
        values=self.booksheet.iter_rows(min_row=self.min_row,min_col=2,max_row=self.max_row,max_col=2)
        title_list=['%s'%i[0].value for i in values] #获取查询网址
        return title_list

    def write_num(self,num_list):#输入值后一列必须有值 
        #fill_colums=3#填入第三列
        #ws = self.booksheet.insert_cols(fill_colums)
        for index,row in enumerate(self.booksheet.rows):
            insert_row_num=self.min_row-1
            num=index-insert_row_num
            if index>=insert_row_num and num<len(num_list):
                row[2].value=num_list[num]
        self.workbook.save(self.path+'/'+'Data_2020_07_13.xlsx')


